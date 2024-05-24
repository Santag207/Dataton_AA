import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

# Algoritmo de asignación
def asignar_horarios(demanda, empleados, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_laboral):
    # Inicializar la estructura para almacenar los horarios
    horarios = {empleado: [''] * len(demanda) for empleado in empleados}
    capacidad_actual = np.zeros(len(demanda))
    
    # Dividir empleados en dos grupos
    mitad = len(empleados) // 2
    empleados_primera_mitad = empleados[:mitad]
    empleados_segunda_mitad = empleados[mitad:]

    # Asignar horarios para todos los empleados escalonadamente desde el inicio del día
    inicio_trabajo = 0  # Iniciar desde el comienzo del día
    for index, empleado in enumerate(empleados):
        for i in range(inicio_trabajo, len(demanda)):
            if horarios[empleado].count('Trabaja') < jornada_laboral:
                if capacidad_actual[i] < demanda['demanda'].iloc[i]:
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                elif horarios[empleado][max(0, i-1)] == 'Trabaja' and (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= minimo_trabajo_continuo:
                    if (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= maximo_trabajo_continuo:
                        horarios[empleado][i] = 'Pausa Activa'
                    else:
                        horarios[empleado][i] = 'Trabaja'
                else:
                    horarios[empleado][i] = 'Trabaja'
        inicio_trabajo += 2  # Escalonar cada 30 minutos (2 franjas de 15 minutos)

    # Asignar almuerzos escalonados para la primera mitad de empleados a partir de las 12:00 PM
    almuerzo_hora = 18  # 12:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_primera_mitad):
        if almuerzo_hora + tiempo_almuerzo <= len(demanda):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
            almuerzo_hora += tiempo_almuerzo

    # Asignar almuerzos escalonados para la segunda mitad de empleados a partir de las 3:00 PM
    almuerzo_hora = 28  # 3:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_segunda_mitad[:-2]):
        if almuerzo_hora + tiempo_almuerzo <= len(demanda):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
            almuerzo_hora += tiempo_almuerzo

    # Asignar almuerzo para el penúltimo empleado de la segunda mitad de 13:00 a 13:45
    if len(empleados_segunda_mitad) > 1:
        penultimo_empleado = empleados_segunda_mitad[-2]
        almuerzo_hora_penultimo = 20  # 13:00 PM en franjas de 15 minutos
        if almuerzo_hora_penultimo + tiempo_almuerzo <= len(demanda):
            horarios[penultimo_empleado][almuerzo_hora_penultimo:almuerzo_hora_penultimo+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo

    # Asignar almuerzo para el último empleado de la segunda mitad de 13:30 a 14:15
    if len(empleados_segunda_mitad) > 0:
        ultimo_empleado = empleados_segunda_mitad[-1]
        almuerzo_hora_ultimo = 22  # 13:30 PM en franjas de 15 minutos
        if almuerzo_hora_ultimo + tiempo_almuerzo <= len(demanda):
            horarios[ultimo_empleado][almuerzo_hora_ultimo:almuerzo_hora_ultimo+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo

    # Asegurar que cada franja horaria tenga al menos un empleado trabajando
    for i in range(len(demanda)):
        if capacidad_actual[i] == 0:
            for empleado in empleados:
                if horarios[empleado][i] == '':
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    break

    return horarios

def run_etapa1():
    # Rutas de archivos
    base_dir = 'optimizacion'
    data_dir = 'xlsx'
    graficas_dir = os.path.join(base_dir, 'graficas')
    
    if not os.path.exists(graficas_dir):
        os.makedirs(graficas_dir)

    demanda_path = os.path.join(data_dir, 'Dataton 2023 Etapa 1.xlsx')
    workers_path = os.path.join(data_dir, 'Dataton 2023 Etapa 1.xlsx')

    # Cargar los Datos
    demanda = pd.read_excel(demanda_path, sheet_name='demand')
    workers = pd.read_excel(workers_path, sheet_name='workers')

    # Mostrar las primeras filas de los datos para revisión
    print(demanda.head())
    print(workers.head())

    # Extraer la hora de la columna 'fecha_hora'
    demanda['hora'] = pd.to_datetime(demanda['fecha_hora']).dt.strftime('%H:%M')

    # Parámetros
    minimo_trabajo_continuo = 4  # franjas de 15 minutos (1 hora)
    maximo_trabajo_continuo = 8  # franjas de 15 minutos (2 horas)
    tiempo_almuerzo = 6  # franjas de 15 minutos (1.5 horas)
    franja_almuerzo_min = 18  # 11:30 AM
    franja_almuerzo_max = 26  # 1:30 PM
    jornada_laboral = 32  # franjas de 15 minutos (8 horas)
    empleados = workers['documento'].unique()
    horarios = asignar_horarios(demanda, empleados, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_laboral)

    def graficar_barras_demanda(demanda, titulo, filename):
        plt.figure(figsize=(10, 6))
    
        # Ancho de las barras
        bar_width = 0.80
        separador_width = 0.001
    
        # Posiciones de las barras
        positions = np.arange(len(demanda))
    
        # Graficar barras de demanda con separadores blancos
        for pos in positions:
            plt.bar(pos, demanda['demanda'][pos], color='gray', width=bar_width)
            plt.bar(pos + bar_width, 0, color='white', width=separador_width)  # Separador blanco
    
        plt.xlabel('Hora del día')
        plt.ylabel('Demanda')
        plt.title(titulo)
        plt.legend(['Demanda'], loc='upper right')
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)
        
        plt.savefig(full_filename)
        plt.close()

    def graficar_demanda_vs_capacidad(demanda, capacidad, titulo, filename):
        plt.figure(figsize=(10, 6))

        # Ancho de las barras
        bar_width = 0.80
        separador_width = 0.001

        # Posiciones de las barras
        positions = np.arange(len(demanda))

        # Graficar barras de demanda con separadores blancos
        for pos in positions:
            plt.bar(pos, demanda['demanda'][pos], color='gray', width=bar_width)
            plt.bar(pos + bar_width, 0, color='white', width=separador_width)  # Separador blanco

        # Graficar la capacidad con un gráfico de escalera
        plt.step(positions, capacidad, label='Capacidad', color='blue', linewidth=3, where='mid')

        plt.xlabel('Hora del día')
        plt.ylabel('Demanda / Capacidad')
        plt.title(titulo)
        plt.legend()
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)
        
        plt.savefig(full_filename)
        plt.close()

    def graficar_diferencia(demanda, capacidad, titulo, filename):
        diferencia = demanda['demanda'] - capacidad
        plt.figure(figsize=(10, 6))
        bar_width = 0.80
        positions = np.arange(len(demanda))

        plt.bar(positions, diferencia, label='Demanda - Capacidad', color=['red' if diff > 0 else 'green' for diff in diferencia], width=bar_width)
        plt.xlabel('Hora del día')
        plt.ylabel('Diferencia')
        plt.title(titulo)
        plt.axhline(0, color='black', linewidth=0.5)
        plt.legend(['Baja Capacidad', 'Sobre Capacidad'], loc='upper right')
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)

        plt.savefig(os.path.join(graficas_dir, filename))
        plt.close()

    # Graficar resultados
    capacidad = np.zeros(len(demanda))
    for i in range(len(demanda)):
        capacidad[i] = sum(1 for horario in horarios.values() if horario[i] == 'Trabaja')

    graficar_barras_demanda(demanda, 'Demanda (Etapa 1)', 'Gpy1_etapa1.png')
    graficar_demanda_vs_capacidad(demanda, capacidad, 'Capacidad vs. Demanda (Etapa 1)', 'Gpy2_etapa1.png')
    graficar_diferencia(demanda, capacidad, 'Demanda - Capacidad (Etapa 1)', 'Gpy3_etapa1.png')

    # Guardar Resultados
    horarios_df = pd.DataFrame.from_dict(horarios, orient='index').transpose()
    horarios_df.index = demanda['fecha_hora']
    horarios_df.reset_index(inplace=True)
    horarios_df.rename(columns={'index': 'fecha_hora'}, inplace=True)
    
    # Ruta del archivo Excel
    excel_filename = os.path.join(base_dir, 'xlsx', 'horarios_optimizados_etapa1.xlsx')
    
    # Eliminar archivo existente si existe
    if os.path.exists(excel_filename):
        os.remove(excel_filename)
    
    horarios_df.to_excel(excel_filename, index=False)

    # Aplicar colores
    wb = load_workbook(excel_filename)
    ws = wb.active

    green_fill = PatternFill(start_color='65FF56', end_color='65FF56', fill_type='solid')
    yellow_fill = PatternFill(start_color='F0FF00', end_color='F0FF00', fill_type='solid')
    blue_fill = PatternFill(start_color='0FA88E', end_color='0FA88E', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Trabaja':
                cell.fill = green_fill
            elif cell.value == 'Almuerza':
                cell.fill = yellow_fill
            elif cell.value == 'Pausa Activa':
                cell.fill = blue_fill

    wb.save(excel_filename)

    def agregar_imagenes_excel(excel_filename, graficas_dir):
        # Cargar el libro de Excel existente
        wb = load_workbook(excel_filename)
    
        # Definir las rutas de las imágenes y los títulos de las hojas
        img_paths = [
            ('Gpy1_etapa1.png', 'Demanda (Etapa 1)'),
            ('Gpy2_etapa1.png', 'Capacidad vs. Demanda (Etapa 1)'),
            ('Gpy3_etapa1.png', 'Demanda - Capacidad (Etapa 1)')
        ]
    
        # Agregar las imágenes a hojas nuevas
        for img_path, title in img_paths:
            ws = wb.create_sheet(title)
            img = Image(os.path.join(graficas_dir, img_path))
            ws.add_image(img, 'A1')
    
        # Guardar el libro de Excel con las imágenes agregadas
        wb.save(excel_filename)

    agregar_imagenes_excel(excel_filename, graficas_dir)

    if __name__ == "__main__":
        run_etapa1()