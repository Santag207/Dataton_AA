import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

# Algoritmo de asignación
def asignar_horarios(demanda, empleados, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_laboral):
    # Inicializar la estructura para almacenar los horarios
    horarios = {empleado: [''] * len(demanda) for empleado in empleados}
    capacidad_actual = np.zeros(len(demanda))
        
    # Dividir empleados en dos grupos
    mitad = len(empleados) // 2
    empleados_primera_mitad = empleados[:mitad]
    empleados_segunda_mitad = empleados[mitad:]

    # Asignar horarios para todos los empleados escalonadamente desde el inicio del día
    inicio_trabajo = 0  # Iniciar desde el comienzo del día
    for index, empleado in enumerate(empleados):
        for i in range(inicio_trabajo, len(demanda)):
            if horarios[empleado].count('Trabaja') < jornada_laboral:
                if capacidad_actual[i] < demanda['demanda'].iloc[i]:
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                elif horarios[empleado][max(0, i-1)] == 'Trabaja' and (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= minimo_trabajo_continuo:
                    if (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= maximo_trabajo_continuo:
                        horarios[empleado][i] = 'Pausa Activa'
                    else:
                        horarios[empleado][i] = 'Trabaja'
                else:
                    horarios[empleado][i] = 'Trabaja'
        inicio_trabajo += 2  # Escalonar cada 30 minutos (2 franjas de 15 minutos)

    # Asignar almuerzos escalonados para la primera mitad de empleados a partir de las 12:00 PM
    almuerzo_hora = 18  # 12:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_primera_mitad):
        if almuerzo_hora + tiempo_almuerzo <= len(demanda):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
            almuerzo_hora += tiempo_almuerzo

    # Asignar almuerzos escalonados para la segunda mitad de empleados a partir de las 3:00 PM
    almuerzo_hora = 28  # 3:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_segunda_mitad[:-2]):
        if almuerzo_hora + tiempo_almuerzo <= len(demanda):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
            almuerzo_hora += tiempo_almuerzo

    # Asignar almuerzo para el penúltimo empleado de la segunda mitad de 13:00 a 13:45
    if len(empleados_segunda_mitad) > 1:
        penultimo_empleado = empleados_segunda_mitad[-2]
        almuerzo_hora_penultimo = 20  # 13:00 PM en franjas de 15 minutos
        if almuerzo_hora_penultimo + tiempo_almuerzo <= len(demanda):
            horarios[penultimo_empleado][almuerzo_hora_penultimo:almuerzo_hora_penultimo+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo

    # Asignar almuerzo para el último empleado de la segunda mitad de 13:30 a 14:15
    if len(empleados_segunda_mitad) > 0:
        ultimo_empleado = empleados_segunda_mitad[-1]
        almuerzo_hora_ultimo = 22  # 13:30 PM en franjas de 15 minutos
        if almuerzo_hora_ultimo + tiempo_almuerzo <= len(demanda):
            horarios[ultimo_empleado][almuerzo_hora_ultimo:almuerzo_hora_ultimo+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo

    # Asegurar que cada franja horaria tenga al menos un empleado trabajando
    for i in range(len(demanda)):
        if capacidad_actual[i] == 0:
            for empleado in empleados:
                if horarios[empleado][i] == '':
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    break

    return horarios

def run_etapa2():
    # Rutas de archivos
    base_dir = 'optimizacion'
    data_dir = 'xlsx'
    graficas_dir = os.path.join(base_dir, 'graficas')
    
    if not os.path.exists(graficas_dir):
        os.makedirs(graficas_dir)

    demanda_path = os.path.join(data_dir, 'Dataton 2023 Etapa 2.xlsx')
    workers_path = os.path.join(data_dir, 'Dataton 2023 Etapa 2.xlsx')

    # Cargar los Datos
    demanda = pd.read_excel(demanda_path, sheet_name='demand')
    workers = pd.read_excel(workers_path, sheet_name='workers')

    # Mostrar las primeras filas de los datos para revisión
    print(demanda.head())
    print(workers.head())

    # Verificar que la columna 'hora' exista, si no, usar otro nombre
    if 'hora' not in demanda.columns:
        if 'fecha_hora' in demanda.columns:
            demanda['hora'] = pd.to_datetime(demanda['fecha_hora']).dt.strftime('%H:%M')
        else:
            print(f"Error: No se encontró la columna 'hora' ni 'fecha_hora' en los datos de demanda.")
            return

    # Parámetros
    minimo_trabajo_continuo = 4  # franjas de 15 minutos (1 hora)
    maximo_trabajo_continuo = 8  # franjas de 15 minutos (2 horas)
    tiempo_almuerzo = 3  # franjas de 15 minutos (45 minutos)
    franja_almuerzo_min = 18  # 11:30 AM
    franja_almuerzo_max = 26  # 1:30 PM
    jornada_completa = 28  # franjas de 15 minutos (7 horas)
    jornada_medio_tiempo = 16  # franjas de 15 minutos (4 horas)
    jornada_sabado_completa = 20  # franjas de 15 minutos (5 horas)
    empleados = workers['documento'].unique()

    # Procesar la demanda día a día
    demanda['fecha'] = pd.to_datetime(demanda['fecha_hora']).dt.date
    dias = demanda['fecha'].unique()
    all_horarios = {empleado: [] for empleado in empleados}

    for dia in dias:
        demanda_dia = demanda[demanda['fecha'] == dia]
        horarios_dia = asignar_horarios(demanda_dia, empleados, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_completa)
        for empleado in empleados:
            all_horarios[empleado] += horarios_dia[empleado]


    # Convertir all_horarios a DataFrame
    horarios_df = pd.DataFrame.from_dict(all_horarios, orient='index').transpose()
    horarios_df.index = demanda['fecha_hora']
    horarios_df.reset_index(inplace=True)
    horarios_df.rename(columns={'index': 'fecha_hora'}, inplace=True)
    
    # Guardar resultados
    excel_filename = os.path.join(base_dir, 'xlsx', 'horarios_optimizados_etapa2.xlsx')
    
    # Eliminar archivo existente si existe
    if os.path.exists(excel_filename):
        os.remove(excel_filename)
    
    horarios_df.to_excel(excel_filename, index=False)

    # Aplicar colores
    wb = load_workbook(excel_filename)
    ws = wb.active

    green_fill = PatternFill(start_color='65FF56', end_color='65FF56', fill_type='solid')
    yellow_fill = PatternFill(start_color='F0FF00', end_color='F0FF00', fill_type='solid')
    blue_fill = PatternFill(start_color='0FA88E', end_color='0FA88E', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Trabaja':
                cell.fill = green_fill
            elif cell.value == 'Almuerza':
                cell.fill = yellow_fill
            elif cell.value == 'Pausa Activa':
                cell.fill = blue_fill

    wb.save(excel_filename)

    def agregar_imagenes_excel(excel_filename, graficas_dir):
        # Cargar el libro de Excel existente
        wb = load_workbook(excel_filename)
    
        # Definir las rutas de las imágenes y los títulos de las hojas
        img_paths = [
            ('Gpy1_etapa2.png', 'Demanda (Etapa 2)'),
            ('Gpy2_etapa2.png', 'Capacidad vs. Demanda (Etapa 2)'),
            ('Gpy3_etapa2.png', 'Demanda - Capacidad (Etapa 2)')
        ]
    
        # Agregar las imágenes a hojas nuevas
        for img_path, title in img_paths:
            ws = wb.create_sheet(title)
            img = Image(os.path.join(graficas_dir, img_path))
            ws.add_image(img, 'A1')
    
        # Guardar el libro de Excel con las imágenes agregadas
        wb.save(excel_filename)

    agregar_imagenes_excel(excel_filename, graficas_dir)

if __name__ == "__main__":
    run_etapa2()


