import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

# Algoritmo de asignación para Etapa 1
def asignar_horarios_etapa1(demanda, empleados, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_laboral):
    # Inicializar la estructura para almacenar los horarios
    horarios = {empleado: [''] * len(demanda) for empleado in empleados}
    capacidad_actual = np.zeros(len(demanda))
    
    # Asignar horarios para todos los empleados escalonadamente desde el inicio del día
    for empleado in empleados:
        trabajo_continuo = 0
        ha_almorzado = False
        for i in range(len(demanda)):
            if trabajo_continuo == maximo_trabajo_continuo:
                if not ha_almorzado and franja_almuerzo_min <= i <= franja_almuerzo_max - tiempo_almuerzo:
                    horarios[empleado][i:i+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
                    ha_almorzado = True
                    trabajo_continuo = 0
                else:
                    horarios[empleado][i] = 'Pausa Activa'
                    trabajo_continuo = 0
            elif trabajo_continuo < maximo_trabajo_continuo:
                if capacidad_actual[i] < demanda['demanda'].iloc[i]:
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    trabajo_continuo += 1
                elif trabajo_continuo >= minimo_trabajo_continuo:
                    if not ha_almorzado and franja_almuerzo_min <= i <= franja_almuerzo_max - tiempo_almuerzo:
                        horarios[empleado][i:i+tiempo_almuerzo] = ['Almuerza'] * tiempo_almuerzo
                        ha_almorzado = True
                        trabajo_continuo = 0
                    else:
                        horarios[empleado][i] = 'Pausa Activa'
                        trabajo_continuo = 0
                else:
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    trabajo_continuo += 1

    # Asegurar que cada franja horaria tenga al menos un empleado trabajando
    for i in range(len(demanda)):
        if capacidad_actual[i] == 0:
            for empleado in empleados:
                if horarios[empleado][i] == '':
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    break

    return horarios

# Algoritmo de asignación para Etapa 2
def asignar_horarios_etapa2(demanda_etapa2, empleados_etapa2, franja_almuerzo_min_etapa2, franja_almuerzo_max_etapa2, minimo_trabajo_continuo_etapa2, maximo_trabajo_continuo_etapa2, tiempo_almuerzo_etapa2, jornada_laboral_etapa2):
    # Inicializar la estructura para almacenar los horarios
    horarios = {empleado: [''] * len(demanda_etapa2) for empleado in empleados_etapa2}
    capacidad_actual = np.zeros(len(demanda_etapa2))
        
    # Dividir empleados en dos grupos
    mitad = len(empleados_etapa2) // 2
    empleados_primera_mitad = empleados_etapa2[:mitad]
    empleados_segunda_mitad = empleados_etapa2[mitad:]

    # Asignar horarios para todos los empleados escalonadamente desde el inicio del día
    inicio_trabajo = 0  # Iniciar desde el comienzo del día
    for index, empleado in enumerate(empleados_etapa2):
        for i in range(inicio_trabajo, len(demanda_etapa2)):
            if horarios[empleado].count('Trabaja') < jornada_laboral_etapa2:
                if capacidad_actual[i] < demanda_etapa2['demanda'].iloc[i]:
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                elif horarios[empleado][max(0, i-1)] == 'Trabaja' and (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= minimo_trabajo_continuo_etapa2:
                    if (i - max(0, horarios[empleado].index('Trabaja')) + 1) >= maximo_trabajo_continuo_etapa2:
                        horarios[empleado][i] = 'Pausa Activa'
                    else:
                        horarios[empleado][i] = 'Trabaja'
                else:
                    horarios[empleado][i] = 'Trabaja'
        inicio_trabajo += 2  # Escalonar cada 30 minutos (2 franjas de 15 minutos)

    # Asignar almuerzos escalonados para la primera mitad de empleados a partir de las 12:00 PM
    almuerzo_hora = 18  # 12:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_primera_mitad):
        if almuerzo_hora + tiempo_almuerzo_etapa2 <= len(demanda_etapa2):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo_etapa2] = ['Almuerza'] * tiempo_almuerzo_etapa2
            almuerzo_hora += tiempo_almuerzo_etapa2

    # Asignar almuerzos escalonados para la segunda mitad de empleados a partir de las 3:00 PM
    almuerzo_hora = 28  # 3:00 PM en franjas de 15 minutos
    for index, empleado in enumerate(empleados_segunda_mitad[:-2]):
        if almuerzo_hora + tiempo_almuerzo_etapa2 <= len(demanda_etapa2):
            horarios[empleado][almuerzo_hora:almuerzo_hora+tiempo_almuerzo_etapa2] = ['Almuerza'] * tiempo_almuerzo_etapa2
            almuerzo_hora += tiempo_almuerzo_etapa2

    # Asignar almuerzo para el penúltimo empleado de la segunda mitad de 13:00 a 13:45
    if len(empleados_segunda_mitad) > 1:
        penultimo_empleado = empleados_segunda_mitad[-2]
        almuerzo_hora_penultimo = 20  # 13:00 PM en franjas de 15 minutos
        if almuerzo_hora_penultimo + tiempo_almuerzo_etapa2 <= len(demanda_etapa2):
            horarios[penultimo_empleado][almuerzo_hora_penultimo:almuerzo_hora_penultimo+tiempo_almuerzo_etapa2] = ['Almuerza'] * tiempo_almuerzo_etapa2

    # Asignar almuerzo para el último empleado de la segunda mitad de 13:30 a 14:15
    if len(empleados_segunda_mitad) > 0:
        ultimo_empleado = empleados_segunda_mitad[-1]
        almuerzo_hora_ultimo = 22  # 13:30 PM en franjas de 15 minutos
        if almuerzo_hora_ultimo + tiempo_almuerzo_etapa2 <= len(demanda_etapa2):
            horarios[ultimo_empleado][almuerzo_hora_ultimo:almuerzo_hora_ultimo+tiempo_almuerzo_etapa2] = ['Almuerza'] * tiempo_almuerzo_etapa2

    # Asegurar que cada franja horaria tenga al menos un empleado trabajando
    for i in range(len(demanda_etapa2)):
        if capacidad_actual[i] == 0:
            for empleado in empleados_etapa2:
                if horarios[empleado][i] == '':
                    horarios[empleado][i] = 'Trabaja'
                    capacidad_actual[i] += 1
                    break

    return horarios

def run_completo():
    # Rutas de archivos
    base_dir = 'optimizacion'
    data_dir = 'xlsx'
    graficas_dir = os.path.join(base_dir, 'graficas')
    
    if not os.path.exists(graficas_dir):
        os.makedirs(graficas_dir)

    demanda_etapa1_path = os.path.join(data_dir, 'Dataton 2023 Etapa 1.xlsx')
    workers_etapa1_path = os.path.join(data_dir, 'Dataton 2023 Etapa 1.xlsx')
    demanda_etapa2_path = os.path.join(data_dir, 'Dataton 2023 Etapa 2.xlsx')
    workers_etapa2_path = os.path.join(data_dir, 'Dataton 2023 Etapa 2.xlsx')

    # Verificar que los archivos existan
    if not os.path.exists(demanda_etapa1_path):
        print(f"Error: El archivo {demanda_etapa1_path} no existe.")
        return
    if not os.path.exists(workers_etapa1_path):
        print(f"Error: El archivo {workers_etapa1_path} no existe.")
        return
    if not os.path.exists(demanda_etapa2_path):
        print(f"Error: El archivo {demanda_etapa2_path} no existe.")
        return
    if not os.path.exists(workers_etapa2_path):
        print(f"Error: El archivo {workers_etapa2_path} no existe.")
        return

    # Cargar los Datos
    demanda_etapa1 = pd.read_excel(demanda_etapa1_path, sheet_name='demand')
    workers_etapa1 = pd.read_excel(workers_etapa1_path, sheet_name='workers')
    demanda_etapa2 = pd.read_excel(demanda_etapa2_path, sheet_name='demand')
    workers_etapa2 = pd.read_excel(workers_etapa2_path, sheet_name='workers')

    # Extraer la hora de la columna 'fecha_hora'
    if 'fecha_hora' in demanda_etapa1.columns:
        demanda_etapa1['hora'] = pd.to_datetime(demanda_etapa1['fecha_hora']).dt.strftime('%H:%M')
    if 'fecha_hora' in demanda_etapa2.columns:
        demanda_etapa2['hora'] = pd.to_datetime(demanda_etapa2['fecha_hora']).dt.strftime('%H:%M')

    # Parámetros comunes
    minimo_trabajo_continuo = 4  # franjas de 15 minutos (1 hora)
    maximo_trabajo_continuo = 8  # franjas de 15 minutos (2 horas)
    tiempo_almuerzo = 6  # franjas de 15 minutos (1.5 horas)
    franja_almuerzo_min = 18  # 11:30 AM
    franja_almuerzo_max = 26  # 1:30 PM

    # Parámetros específicos para Etapa 1
    jornada_laboral_etapa1 = 32  # franjas de 15 minutos (8 horas)
    empleados_etapa1 = workers_etapa1['documento'].unique()

    # Llamada a la función de asignación de horarios para etapa 1
    horarios_etapa1 = asignar_horarios_etapa1(demanda_etapa1, empleados_etapa1, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_laboral_etapa1)

    # Parámetros específicos para Etapa 2
    jornada_completa = 28  # franjas de 15 minutos (7 horas)
    jornada_medio_tiempo = 16  # franjas de 15 minutos (4 horas)
    jornada_sabado_completa = 20  # franjas de 15 minutos (5 horas)
    empleados_etapa2 = workers_etapa2['documento'].unique()

    # Llamada a la función de asignación de horarios para etapa 2
   # Procesar la demanda día a día para Etapa 2
    demanda_etapa2['fecha'] = pd.to_datetime(demanda_etapa2['fecha_hora']).dt.date
    dias_etapa2 = demanda_etapa2['fecha'].unique()
    all_horarios_etapa2 = {empleado: [] for empleado in empleados_etapa2}

    for dia in dias_etapa2:
        demanda_dia_etapa2 = demanda_etapa2[demanda_etapa2['fecha'] == dia]
        horarios_dia_etapa2 = asignar_horarios_etapa2(demanda_dia_etapa2, empleados_etapa2, franja_almuerzo_min, franja_almuerzo_max, minimo_trabajo_continuo, maximo_trabajo_continuo, tiempo_almuerzo, jornada_completa)
        for empleado in empleados_etapa2:
            all_horarios_etapa2[empleado] += horarios_dia_etapa2[empleado]

    # Mostrar las primeras filas de los datos para revisión
    print(demanda_etapa1.head())
    print(workers_etapa1.head())

    # Mostrar las primeras filas de los datos para revisión
    print(demanda_etapa2.head())
    print(workers_etapa2.head())

    # Visualización y Validación
    def graficar_barras_demanda(demanda, titulo, filename):
        plt.figure(figsize=(10, 6))
    
        # Ancho de las barras
        bar_width = 0.80
        separador_width = 0.001
    
        # Posiciones de las barras
        positions = np.arange(len(demanda))
    
        # Graficar barras de demanda con separadores blancos
        for pos in positions:
            plt.bar(pos, demanda['demanda'][pos], color='gray', width=bar_width)
            plt.bar(pos + bar_width, 0, color='white', width=separador_width)  # Separador blanco
    
        plt.xlabel('Hora del día')
        plt.ylabel('Demanda')
        plt.title(titulo)
        plt.legend(['Demanda'], loc='upper right')
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)
        
        plt.savefig(full_filename)
        plt.close()

    def graficar_demanda_vs_capacidad(demanda, capacidad, titulo, filename):
        plt.figure(figsize=(10, 6))

        # Ancho de las barras
        bar_width = 0.80
        separador_width = 0.001

        # Posiciones de las barras
        positions = np.arange(len(demanda))

        # Graficar barras de demanda con separadores blancos
        for pos in positions:
            plt.bar(pos, demanda['demanda'][pos], color='gray', width=bar_width)
            plt.bar(pos + bar_width, 0, color='white', width=separador_width)  # Separador blanco

        # Graficar la capacidad con un gráfico de escalera
        plt.step(positions, capacidad, label='Capacidad', color='blue', linewidth=3, where='mid')

        plt.xlabel('Hora del día')
        plt.ylabel('Demanda / Capacidad')
        plt.title(titulo)
        plt.legend()
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)
        
        plt.savefig(full_filename)
        plt.close()

    def graficar_diferencia(demanda, capacidad, titulo, filename):
        diferencia = demanda['demanda'] - capacidad
        plt.figure(figsize=(10, 6))
        bar_width = 0.80
        positions = np.arange(len(demanda))

        plt.bar(positions, diferencia, label='Demanda - Capacidad', color=['red' if diff > 0 else 'green' for diff in diferencia], width=bar_width)
        plt.xlabel('Hora del día')
        plt.ylabel('Diferencia')
        plt.title(titulo)
        plt.axhline(0, color='black', linewidth=0.5)
        plt.legend(['Baja Capacidad', 'Sobre Capacidad'], loc='upper right')
        plt.xticks(ticks=positions, labels=demanda['hora'], rotation=90, ha='right', fontsize=8)
        plt.tight_layout()
        
        # Eliminar archivo existente si existe
        full_filename = os.path.join(graficas_dir, filename)
        if os.path.exists(full_filename):
            os.remove(full_filename)

        plt.savefig(os.path.join(graficas_dir, filename))
        plt.close()

    # Graficar resultados para Etapa 1
    capacidad = np.zeros(len(demanda_etapa1))
    for i in range(len(demanda_etapa1)):
        capacidad[i] = sum(1 for horario in horarios_etapa1.values() if horario[i] == 'Trabaja')

    graficar_barras_demanda(demanda_etapa1, 'Demanda (Etapa 1)', 'Gpy1_etapa1.png')
    graficar_demanda_vs_capacidad(demanda_etapa1, capacidad, 'Capacidad vs. Demanda (Etapa 1)', 'Gpy2_etapa1.png')
    graficar_diferencia(demanda_etapa1, capacidad, 'Demanda - Capacidad (Etapa 1)', 'Gpy3_etapa1.png')

    # Graficar resultados para Etapa 2
    capacidad = np.zeros(len(demanda_etapa2))
    for i in range(len(demanda_etapa2)):
        capacidad[i] = sum(1 for horario in all_horarios_etapa2.values() if horario[i] == 'Trabaja')

    graficar_barras_demanda(demanda_etapa2, 'Demanda (Etapa 2)', 'Gpy1_etapa2.png')
    graficar_demanda_vs_capacidad(demanda_etapa2, capacidad, 'Capacidad vs. Demanda (Etapa 2)', 'Gpy2_etapa2.png')
    graficar_diferencia(demanda_etapa2, capacidad, 'Demanda - Capacidad (Etapa 2)', 'Gpy3_etapa2.png')

    # Guardar Resultados para Etapa 1
    horarios_df_etapa1 = pd.DataFrame.from_dict(horarios_etapa1, orient='index').transpose()
    horarios_df_etapa1.index = demanda_etapa1['fecha_hora']
    horarios_df_etapa1.reset_index(inplace=True)
    horarios_df_etapa1.rename(columns={'index': 'fecha_hora'}, inplace=True)

    # Ruta del archivo Excel
    excel_filename = os.path.join(base_dir, 'xlsx', 'horarios_optimizados_etapa1.xlsx')
    
    # Eliminar archivo existente si existe
    if os.path.exists(excel_filename):
        os.remove(excel_filename)
    
    horarios_df_etapa1.to_excel(excel_filename, index=False)

    # Aplicar colores
    wb = load_workbook(excel_filename)
    ws = wb.active

    green_fill = PatternFill(start_color='65FF56', end_color='65FF56', fill_type='solid')
    yellow_fill = PatternFill(start_color='F0FF00', end_color='F0FF00', fill_type='solid')
    blue_fill = PatternFill(start_color='0FA88E', end_color='0FA88E', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Trabaja':
                cell.fill = green_fill
            elif cell.value == 'Almuerza':
                cell.fill = yellow_fill
            elif cell.value == 'Pausa Activa':
                cell.fill = blue_fill

    wb.save(excel_filename)

    def agregar_imagenes_excel(excel_filename, graficas_dir):
        # Cargar el libro de Excel existente
        wb = load_workbook(excel_filename)
    
        # Definir las rutas de las imágenes y los títulos de las hojas
        img_paths = [
            ('Gpy1_etapa1.png', 'Demanda (Etapa 1)'),
            ('Gpy2_etapa1.png', 'Capacidad vs. Demanda (Etapa 1)'),
            ('Gpy3_etapa1.png', 'Demanda - Capacidad (Etapa 1)')
        ]
    
        # Agregar las imágenes a hojas nuevas
        for img_path, title in img_paths:
            ws = wb.create_sheet(title)
            img = Image(os.path.join(graficas_dir, img_path))
            ws.add_image(img, 'A1')
    
        # Guardar el libro de Excel con las imágenes agregadas
        wb.save(excel_filename)

    agregar_imagenes_excel(excel_filename, graficas_dir)

    # Guardar Resultados para Etapa 2
    horarios_df_etapa2 = pd.DataFrame.from_dict(all_horarios_etapa2, orient='index').transpose()
    horarios_df_etapa2.index = demanda_etapa2['fecha_hora']
    horarios_df_etapa2.reset_index(inplace=True)
    horarios_df_etapa2.rename(columns={'index': 'fecha_hora'}, inplace=True)

    # Ruta del archivo Excel
    excel_filename = os.path.join(base_dir, 'xlsx', 'horarios_optimizados_etapa2.xlsx')
    
    # Eliminar archivo existente si existe
    if os.path.exists(excel_filename):
        os.remove(excel_filename)

    horarios_df_etapa2.to_excel(excel_filename, index=False)

    # Aplicar colores
    wb = load_workbook(excel_filename)
    ws = wb.active

    green_fill = PatternFill(start_color='65FF56', end_color='65FF56', fill_type='solid')
    yellow_fill = PatternFill(start_color='F0FF00', end_color='F0FF00', fill_type='solid')
    blue_fill = PatternFill(start_color='0FA88E', end_color='0FA88E', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Trabaja':
                cell.fill = green_fill
            elif cell.value == 'Almuerza':
                cell.fill = yellow_fill
            elif cell.value == 'Pausa Activa':
                cell.fill = blue_fill

    wb.save(excel_filename)

    def agregar_imagenes_excel(excel_filename, graficas_dir):
        # Cargar el libro de Excel existente
        wb = load_workbook(excel_filename)
    
        # Definir las rutas de las imágenes y los títulos de las hojas
        img_paths = [
            ('Gpy1_etapa2.png', 'Demanda (Etapa 2)'),
            ('Gpy2_etapa2.png', 'Capacidad vs. Demanda (Etapa 2)'),
            ('Gpy3_etapa2.png', 'Demanda - Capacidad (Etapa 2)')
        ]
    
        # Agregar las imágenes a hojas nuevas
        for img_path, title in img_paths:
            ws = wb.create_sheet(title)
            img = Image(os.path.join(graficas_dir, img_path))
            ws.add_image(img, 'A1')
    
        # Guardar el libro de Excel con las imágenes agregadas
        wb.save(excel_filename)

    agregar_imagenes_excel(excel_filename, graficas_dir)

if __name__ == "__main__":
    run_completo()